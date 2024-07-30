# -*- coding: utf-8 -*-
"""
Created on Thu Jul 11 13:00:12 2024
This script is used to calibrate resident mode choice bike and walk alternative specific constants.
The script is run for each iteration of calibration. To start, there must be a copy of the existing UEC in "file_path", 
and a copy of the same UEC with a different name in "file_path2". File path 1 will only be read, and file_path2 will be
updated with new ASCs

Set the iteration number and the mdl_path. The model outputs will be read and summarized and the adjustments will be calculated
and applied to the appropriate ASCs and written into file_path2

Once the script is complete, replace the model UEC with the updated UEC and run the model again. Further iteratsions are needed
update the file_path, the iteration number, and the model directory and run again. 

A log is printed and appended each time the script is run to show the current mode share and the current targets.

@author: hannah.carson hannah.carson@rsginc.com
"""



import pandas as pd
from xlrd import open_workbook
from openpyxl import load_workbook
import xlwt
import xlrd
from xlutils.copy import copy
import os
import numpy as np

os.chdir(r'S:\Projects\CA\TRPA\tasks\BikeCalibration')
# path to UEC file with starting coefficients
file_path = r"S:\Projects\CA\TRPA\tasks\BikeCalibration\output\Mode_Choice.xls"
#path to copy of current UEC where updated coefficents will be saved
file_path2 = r"S:\Projects\CA\TRPA\tasks\BikeCalibration\output\Mode_Choice_1.xls"

mdl_path = r'S:\Projects\CA\TRPA\Model\3_Updated_Model_compiled_java\TahoeModel_Release1\scenarios\2018'
iteration = 1

#targets
work_bike_target = 0.005
work_walk_target = 0.023

othr_bike_target = 0.012
othr_walk_target = 0.069

# Function to read Excel sheets with xlrd
def read_excel_sheets(file_path):
    workbook = open_workbook(file_path)
    sheetnames = workbook.sheet_names()
    dfs = []

    for sheetname in sheetnames[2:]:
        if 'OD' in sheetname:
            continue
        sheet = workbook.sheet_by_name(sheetname)
        # Read the sheet into a pandas dataframe
        debug = False
        if sheetname == 'atworkOD':
            debug = True
        data = {sheet.cell_value(2, col): [sheet.cell_value(row, col) for row in range(3, sheet.nrows)] for col in range(sheet.ncols)}
        df = pd.DataFrame(data)

        # Filter columns: keep only 'No', 'alt5', and 'alt6' columns
        columns_to_keep = ['No','Formula_for_variable'] + [col for col in df.columns if col.lower() in ['description','alt5', 'alt6']]
        df_filtered = df[columns_to_keep]
        
        # Add a sheet identifier
        df_filtered['sheet'] = sheetname
        
        # Append to the list of dataframes
        dfs.append(df_filtered)

    return pd.concat(dfs, ignore_index=True)

# Read all sheets into a single dataframe
combined_df = read_excel_sheets(file_path)
combined_df = combined_df[combined_df['No'].notna() & (combined_df['No'] != '')]
combined_df = combined_df[combined_df['Formula_for_variable']!='@@ODUtilModeAlt']
combined_df = combined_df[(combined_df['Alt5'].notna() & (combined_df['Alt5'] != '') & (combined_df['Alt5'] != 0)) | (combined_df['Alt6'].notna() & (combined_df['Alt6'] != '') & (combined_df['Alt6'] != 0))]
# Manipulate your dataframe here
# Example: Add a new column with manipulated values

trips = pd.read_csv(os.path.join(mdl_path, 'outputs_summer','trip_file.csv'))
trips = trips[trips.partyType == 'resident']
work_trips = trips[trips.tripType == 'Work']
othr_trips = trips[trips.tripType != 'Work']

work_bike_share = len(work_trips[work_trips['mode'] == 'bike'])/len(work_trips)
work_walk_share = len(work_trips[work_trips['mode'] == 'walk'])/len(work_trips)

othr_bike_share = len(othr_trips[othr_trips['mode'] == 'bike'])/len(othr_trips)
othr_walk_share = len(othr_trips[othr_trips['mode'] == 'walk'])/len(othr_trips)

with open(r'output\calib_log.txt','a') as f:
    f.write("Iteration: {}\n".format(iteration))
    f.write("Work Bike Mode Share: {}\n".format(work_bike_share))
    f.write("Work Bike Target Share: {}\n".format(work_bike_target))
    f.write("Work Bike Diff Share: {}\n".format(work_bike_share - work_bike_target))
    f.write("\n\n\n")
    f.write("Other Bike Mode Share: {}\n".format(othr_bike_share))
    f.write("Other Bike Target Share: {}\n".format(othr_bike_target))
    f.write("Other Bike Diff Share: {}\n".format(othr_bike_share - othr_bike_target))    
    f.write("\n\n\n")
    f.write("Work Walk Mode Share: {}\n".format(work_walk_share))
    f.write("Work Walk Target Share: {}\n".format(work_walk_target))
    f.write("Work Walk Diff Share: {}\n".format(work_walk_share - work_walk_target))
    f.write("\n\n\n")
    f.write("Other Walk Mode Share: {}\n".format(othr_walk_share))
    f.write("Other Walk Target Share: {}\n".format(othr_walk_target))
    f.write("Other Walk Diff Share: {}\n".format(othr_walk_share - othr_walk_target))    
    f.write("\n\n\n")
    
f.close() 

work_walk_adj = np.log(work_walk_target/work_walk_share)
work_bike_adj = np.log(work_bike_target/work_bike_share)
othr_walk_adj = np.log(othr_walk_target/othr_walk_share)
othr_bike_adj = np.log(othr_bike_target/othr_bike_share)

combined_df['is_work'] = combined_df['sheet'].map(lambda n: n in ['work','workOD'])
combined_df.loc[combined_df['is_work'], 'Alt5'] = combined_df.loc[combined_df['is_work'], 'Alt5'].apply(lambda x: x + work_walk_adj if isinstance(x, float) and x != -999 and x != 1 else x)
combined_df.loc[~combined_df['is_work'],'Alt5'] = combined_df.loc[~combined_df['is_work'], 'Alt5'].apply(lambda x: x + othr_walk_adj if isinstance(x, float) and x != -999 and x != 1 else x)
combined_df.loc[combined_df['is_work'], 'Alt6'] = combined_df.loc[combined_df['is_work'], 'Alt6'].apply(lambda x: x + work_bike_adj if isinstance(x, float) and x != -999 and x != 1 else x)
combined_df.loc[~combined_df['is_work'],'Alt6'] = combined_df.loc[~combined_df['is_work'], 'Alt6'].apply(lambda x: x + othr_bike_adj if isinstance(x, float) and x != -999 and x != 1 else x)


# Function to update the workbook with the new values
def update_workbook(file_path, combined_df):
    # Open the existing workbook with xlrd
    rb = xlrd.open_workbook(file_path, formatting_info=True)
    wb = copy(rb)
    
    for sheetname in rb.sheet_names()[2:]:
        # Load the dataframe for the current sheet
        df_sheet = combined_df[combined_df['sheet'] == sheetname]
        
        # Load the sheet
        sheet = rb.sheet_by_name(sheetname)
        ws = wb.get_sheet(sheetname)
        
        # Create a mapping of column headers to their index positions
        header_to_index = {sheet.cell_value(2, col): col for col in range(sheet.ncols)}
        
        # Iterate through the rows of the Excel sheet
        for excel_row in range(3, sheet.nrows):
            excel_No = sheet.cell_value(excel_row, header_to_index['No'])
            
            # Find the matching row in the dataframe based on the No
            matching_row = df_sheet[df_sheet['No'] == excel_No]
            if not matching_row.empty:
                matching_row = matching_row.iloc[0]
                for col_name in ['Alt5', 'Alt6']:
                    if col_name in matching_row and matching_row[col_name] != '':
                        col_num = header_to_index[col_name]
                        current_value = sheet.cell_value(excel_row, col_num)
                        if current_value:  # Only update if the current cell value is not empty or blank
                            ws.write(excel_row, col_num, matching_row[col_name])
    
    # Save the workbook
    wb.save(file_path)   
update_workbook(file_path2, combined_df)

print("Excel workbook has been updated successfully.")
